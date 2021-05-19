import openpyxl


class HomePageData:

    form_data = [{'firstname': 'Yevhen', 'lastname': 'Domin',
                  'email': 'evhen@gmail.com', 'gender': 'Male'},
                 {'firstname': 'Alina', 'lastname': 'Domina',
                  'email': 'a.domina@gmail.com', 'gender': 'Female'}]

    @staticmethod
    def get_test_data_excel(test_name):
        data_dict = {}
        book = openpyxl.load_workbook(r'D:\pytest_selenium_framework\TestData\HomePageData.xlsx')
        sheet = book.active
        cell = sheet.cell(row=2, column=2)

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_name:
                for j in range(2, sheet.max_column + 1):
                    if i == 1:
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i + 1, column=j).value
                    else:
                        data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [data_dict]
